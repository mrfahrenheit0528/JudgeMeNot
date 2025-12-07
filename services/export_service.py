import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.utils import ImageReader

class ExportService:
    def generate_excel(self, filepath, event_name, title, data_matrix, mode="segment"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tabulation"
        
        ws['A1'] = event_name; ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = title; ws['A2'].font = Font(size=14, bold=True)
        
        row_num = 4
        cols = data_matrix.get('judges', []) if mode == 'segment' else data_matrix.get('segments', [])
        
        def write_gender_table(gender_name, rows):
            nonlocal row_num
            ws.cell(row=row_num, column=1, value=f"{gender_name} RANKING").font = Font(bold=True)
            row_num += 1
            headers = ["Rank", "#", "Candidate"] + cols + ["Total"]
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=h)
                cell.font = Font(bold=True)
                cell.border = Border(bottom=Side(style='thin'))
            row_num += 1
            for r in rows:
                scores = r['scores'] if mode == 'segment' else r['segment_scores']
                row_data = [r['rank'], r['number'], r['name']] + scores + [r['total']]
                for col_idx, val in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_idx, value=val)
                row_num += 1
            row_num += 2

        write_gender_table("MALE", data_matrix['Male'])
        write_gender_table("FEMALE", data_matrix['Female'])
        
        wb.save(filepath)
        return True

    def generate_pdf(self, filepath, event_name, title, data_matrix, mode="segment"):
        # Custom Paper Size: 8.5" x 13" (Folio/Long Bond Paper)
        FOLIO_SIZE = (8.5 * inch, 13 * inch)
        
        # Define Header Function
        def add_header(canvas, doc):
            canvas.saveState()
            page_width, page_height = FOLIO_SIZE
            
            # --- 1. HEADER IMAGE ---
            # We assume the banner image is saved as 'assets/header.png'
            header_path = "assets/header.png"
            
            if os.path.exists(header_path):
                # Draw Image full width at top
                # Aspect ratio preservation logic can be added, but stretching to header often looks best for banners
                header_height = 1.5 * inch 
                canvas.drawImage(header_path, 0, page_height - header_height, width=page_width, height=header_height, mask='auto')
            else:
                # Fallback Text if image missing
                canvas.setFont('Helvetica-Bold', 10)
                canvas.drawCentredString(page_width / 2, page_height - 0.5 * inch, "Camarines Sur Polytechnic Colleges")
                canvas.setFont('Helvetica', 10)
                canvas.drawCentredString(page_width / 2, page_height - 0.65 * inch, "College of Computer Studies | Junior Philippine Computer Society")
                canvas.drawCentredString(page_width / 2, page_height - 0.8 * inch, "CSPC Chapter")
                
                # Draw Logo Placeholder if missing
                canvas.rect(0.5*inch, page_height - 1.2*inch, 0.8*inch, 0.8*inch)
                canvas.drawString(0.6*inch, page_height - 0.8*inch, "LOGO")

            # --- 2. FOOTER (Optional Page Number) ---
            canvas.setFont('Helvetica', 9)
            canvas.drawString(0.5 * inch, 0.5 * inch, f"Page {doc.page}")
            canvas.drawRightString(page_width - 0.5 * inch, 0.5 * inch, "System Generated Report")
            
            canvas.restoreState()

        # Doc Template
        doc = SimpleDocTemplate(
            filepath, 
            pagesize=FOLIO_SIZE,
            topMargin=2.0 * inch, # Space for the Header Image
            leftMargin=0.5 * inch,
            rightMargin=0.5 * inch,
            bottomMargin=1.0 * inch
        )
        
        elements = []
        styles = getSampleStyleSheet()
        # Custom styles to center titles
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], alignment=TA_CENTER, spaceAfter=5)
        sub_style = ParagraphStyle('SubStyle', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=20)
        
        # Determine Columns
        cols = data_matrix.get('judges', []) if mode == 'segment' else data_matrix.get('segments', [])
        
        # Helper to build table
        def build_gender_section(gender_name, rows):
            # 1. Title Outside Table
            elements.append(Paragraph(f"{event_name}", title_style))
            elements.append(Paragraph(f"{title} - {gender_name}", sub_style))
            
            # 2. Table Data
            headers = ["Rank", "#", "Candidate"] + cols + ["Total"]
            table_data = [headers]
            
            for r in rows:
                scores = r['scores'] if mode == 'segment' else r['segment_scores']
                row = [str(r['rank']), str(r['number']), r['name']] + [str(s) for s in scores] + [str(r['total'])]
                table_data.append(row)
            
            # 3. Create Table
            col_widths = [0.6*inch, 0.6*inch, 2.5*inch] # Fixed widths for Rank, #, Name
            # Distribute remaining width to score columns
            remaining_width = 7.5*inch - sum(col_widths) - 0.8*inch # 0.8 for Total
            score_col_w = remaining_width / max(len(cols), 1)
            
            final_col_widths = col_widths + [score_col_w]*len(cols) + [0.8*inch]

            t = Table(table_data, colWidths=final_col_widths)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.6)), # Blue Header
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]) # Alternating rows
            ]))
            elements.append(t)
            
            # 4. Signatories
            # REMOVED: if mode == 'segment': check. Now shows for both modes.
            if True: 
                elements.append(Spacer(1, 40))
                elements.append(Paragraph("Certified by:", styles['Normal']))
                elements.append(Spacer(1, 40))
                
                judges_list = data_matrix.get('judges', [])
                sig_data = []
                row = []
                for j in judges_list:
                    # Create a signature line block
                    sig_block = f"_______________________\n{j}\n(Judge)"
                    row.append(sig_block)
                    if len(row) >= 3:
                        sig_data.append(row)
                        row = []
                if row: sig_data.append(row)
                
                if sig_data:
                    sig_table = Table(sig_data)
                    sig_table.setStyle(TableStyle([
                        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                        ('VALIGN', (0,0), (-1,-1), 'TOP'),
                        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
                        ('LEFTPADDING', (0,0), (-1,-1), 20),
                        ('RIGHTPADDING', (0,0), (-1,-1), 20),
                        ('BOTTOMPADDING', (0,0), (-1,-1), 30),
                    ]))
                    elements.append(sig_table)
            
            # 5. Page Break for next gender
            elements.append(PageBreak())

        # Generate Pages
        build_gender_section("MALE RANKING", data_matrix['Male'])
        build_gender_section("FEMALE RANKING", data_matrix['Female'])
        
        # Build PDF with Header Callback
        doc.build(elements, onFirstPage=add_header, onLaterPages=add_header)
        return True